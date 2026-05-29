#!/usr/bin/env python3
"""
Build bu_dashboard_data.json from the source Excel workbook on SharePoint.

Pipeline:
  1. Acquire an app-only Microsoft Graph token (client credentials).
  2. Download the source .xlsx from the PCMA Team site (addressed by stable
     item id, so moving the file between folders does not break this).
  3. Parse it with openpyxl, grouping subgroups by the FILL COLOR of their
     header row in column A into the four top-level groups.
  4. Write bu_dashboard_data.json (the GitHub Action commits it if changed).

The dashboard groups business units by header-row fill color (commit 91a6ee0,
"Updated groupings by cell color not text"). openpyxl is used specifically
because it can read cell fill colors, which the Power Automate / Graph Excel
"list rows" APIs cannot.

Workbook layout (sheet "Business Unit"), confirmed against the live file:
  Col A  group name + color key (only populated on subgroup header rows)
  Col B  subgroup name (header rows) / line-item name (item rows)
  Col C  2026 goal      Col D  Feb YTD
  Col E  Target to Goal (remaining)   Col F  STLY
  Cols G-J (source / manual-api / program / GL codes) are ignored.
  Rows after the last group (Subtotal, Total, Balance Check, ...) are ignored
  because their col-A fill color is not in COLOR_MAP.

Local testing without Graph: set BU_LOCAL_XLSX=/path/to/file.xlsx to parse a
local copy and skip authentication.
"""

import io
import json
import os
import sys
from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

# --------------------------------------------------------------------------
# CONFIG  (file-specific values, confirmed against the live workbook)
# --------------------------------------------------------------------------

GRAPH = "https://graph.microsoft.com/v1.0"

# PCMA Team site document library (stable drive id) and the workbook's stable
# item id. Addressing by item id means the file can be moved between folders
# without breaking this pipeline.
DRIVE_ID = "b!_609OR8fLU6wjK_Inp1_4O2-L5JDgrxHumtyBqVrpD8I85ePiYNhQITFvKReaukz"
ITEM_ID = "01WUVNJBVHHYXW4LBSZRGYE3CTJMQBZ33J"  # 2026 Business Unit Numbers.xlsx

SHEET_NAME = "Business Unit"

# Output file (repo root, one level up from this script).
OUTPUT_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "bu_dashboard_data.json",
)

# Columns (1-based).
COL_GROUP = 1      # A: group name; its fill color is the grouping key
COL_LABEL = 2      # B: subgroup name (header rows) / item name (item rows)
COL_GOAL = 3       # C: 2026 goal
COL_YTD = 4        # D: Feb YTD
COL_REMAINING = 5  # E: Target to Goal (remaining); recomputed if blank
COL_STLY = 6       # F: STLY
HEADER_ROWS = 1    # number of header rows to skip at the top

# Fill-color token -> (group_id, group_header). A token comes from fill_token()
# below: "RGB:<aarrggbb>" for explicit fills, "THEME:<idx>:<tint>" for theme
# fills. Products / AI / Initiative share one fill, so they fold into one group.
COLOR_MAP = {
    "THEME:3:0.75": ("enterprise", "Enterprise"),
    "THEME:9:0.8": ("business_unit", "Business Unit"),
    "RGB:FFFFFFCC": ("events", "Events"),
    "THEME:8:0.8": ("products_ai_initiative", "Products, AI, Initiative"),
}

# Deterministic output order regardless of workbook row order.
GROUP_ORDER = ["enterprise", "business_unit", "events", "products_ai_initiative"]

CHICAGO = ZoneInfo("America/Chicago")

# --------------------------------------------------------------------------
# Graph auth + download (skipped when BU_LOCAL_XLSX is set)
# --------------------------------------------------------------------------


def get_token() -> str:
    import msal

    tenant = os.environ["AZURE_TENANT_ID"]
    client_id = os.environ["AZURE_CLIENT_ID"]
    secret = os.environ["AZURE_CLIENT_SECRET"]
    app = msal.ConfidentialClientApplication(
        client_id,
        authority=f"https://login.microsoftonline.com/{tenant}",
        client_credential=secret,
    )
    result = app.acquire_token_for_client(
        scopes=["https://graph.microsoft.com/.default"]
    )
    if "access_token" not in result:
        raise RuntimeError(
            f"Token request failed: {result.get('error')} "
            f"{result.get('error_description')}"
        )
    return result["access_token"]


def download_workbook() -> bytes:
    import requests

    token = get_token()
    url = f"{GRAPH}/drives/{DRIVE_ID}/items/{ITEM_ID}/content"
    resp = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=120)
    resp.raise_for_status()
    return resp.content


# --------------------------------------------------------------------------
# Parsing
# --------------------------------------------------------------------------


def fill_token(cell) -> str:
    """Normalize a cell's fill into a stable, comparable token."""
    fill = cell.fill
    if fill is None or fill.patternType is None:
        return ""
    fg = fill.fgColor
    if fg is None:
        return ""
    if getattr(fg, "type", None) == "rgb" and fg.rgb and fg.rgb != "00000000":
        return f"RGB:{str(fg.rgb).upper()}"
    if getattr(fg, "type", None) == "theme":
        tint = round(float(fg.tint or 0.0), 3)
        return f"THEME:{fg.theme}:{tint}"
    return ""


def num(value):
    """Coerce an Excel cell value to a number or None."""
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return value
    s = str(value).strip().replace(",", "").replace("$", "")
    if s in ("", "-", "N/A", "n/a"):
        return None
    try:
        return float(s) if "." in s else int(s)
    except ValueError:
        return None


def cell_text(cell):
    v = cell.value
    return None if v is None else str(v).strip()


def parse(content: bytes) -> list:
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    groups = {}
    current_group = None
    current_sub = None

    for r in range(HEADER_ROWS + 1, ws.max_row + 1):
        group_cell = ws.cell(row=r, column=COL_GROUP)
        group_val = cell_text(group_cell)

        if group_val:
            # Column A has text: either a group/subgroup header (known fill
            # color) or a trailing summary row (unknown fill -> stop a group).
            token = fill_token(group_cell)
            if token in COLOR_MAP:
                gid, header = COLOR_MAP[token]
                current_group = groups.setdefault(
                    gid, {"id": gid, "header": header, "subgroups": []}
                )
                current_sub = {"name": cell_text(ws.cell(row=r, column=COL_LABEL)) or "", "items": []}
                current_group["subgroups"].append(current_sub)
            else:
                current_group = None
                current_sub = None
            continue

        # Column A empty: an item row under the current subgroup.
        if current_sub is None:
            continue
        name = cell_text(ws.cell(row=r, column=COL_LABEL))
        if not name:
            continue
        goal = num(ws.cell(row=r, column=COL_GOAL).value)
        ytd = num(ws.cell(row=r, column=COL_YTD).value)
        remaining = num(ws.cell(row=r, column=COL_REMAINING).value)
        stly = num(ws.cell(row=r, column=COL_STLY).value)
        if goal is None and ytd is None:
            continue  # placeholder line with no numbers (e.g. ELF)
        if remaining is None and goal is not None:
            remaining = goal - (ytd or 0)
        current_sub["items"].append(
            {"name": name, "goal": goal, "ytd": ytd, "remaining": remaining, "stly": stly}
        )

    ordered = []
    for gid in GROUP_ORDER:
        g = groups.get(gid)
        if not g:
            continue
        g["subgroups"] = [s for s in g["subgroups"] if s["items"]]
        if g["subgroups"]:
            ordered.append(g)
    return ordered


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------


def main():
    local = os.environ.get("BU_LOCAL_XLSX")
    if local:
        with open(local, "rb") as f:
            content = f.read()
    else:
        content = download_workbook()

    color_groups = parse(content)
    if not color_groups:
        print("ERROR: parser produced no groups; refusing to overwrite output.")
        sys.exit(1)

    payload = {
        "generated_at": datetime.now(CHICAGO).strftime("%b %d, %Y %H:%M"),
        "color_groups": color_groups,
    }
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)
        f.write("\n")

    subs = sum(len(g["subgroups"]) for g in color_groups)
    items = sum(len(s["items"]) for g in color_groups for s in g["subgroups"])
    print(f"Wrote {OUTPUT_PATH}: {len(color_groups)} groups, {subs} subgroups, {items} items.")


if __name__ == "__main__":
    main()
